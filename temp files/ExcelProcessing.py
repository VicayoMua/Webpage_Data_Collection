from openpyxl import load_workbook as load_excel_workbook, Workbook as ExcelWorkbook  # other files may need it, so keep it!!!
from openpyxl.worksheet.worksheet import Worksheet as ExcelWorksheet
from openpyxl.cell import Cell as ExcelCell  # other files may need it, so keep it!!!
from openpyxl.styles import Font as ExcelFont, Alignment as ExcelAlignment, PatternFill as ExcelPatternFill  # other files may need it, so keep it!!!

# To import from this file, please use this interface:
# from ExcelProcessing import (
#     load_excel_workbook,
#     ExcelWorkbook,
#     ExcelWorksheet,
#     ExcelCell,
#     ExcelFont,
#     ExcelAlignment,
#     ExcelPatternFill
# )

def ask_for_excel_filepath(query: str) -> str:
    excel_filepath = input(query)
    if not excel_filepath.endswith(".xlsx"):
        excel_filepath += ".xlsx"
    return excel_filepath


def get_worksheet_data_by_axis(
        data_container: set | list,
        clear_data_container: bool,
        sheet: ExcelWorksheet,
        axis: str,
        value_former_transformer=lambda v: v,
        value_filter=lambda v: True,
        value_post_transformer=lambda v: v
) -> None:
    if isinstance(data_container, set):
        if clear_data_container:
            data_container.clear()
        for cell in sheet[axis]:
            val = value_former_transformer(cell.value)
            if value_filter(val):
                data_container.add(value_post_transformer(val))
    if isinstance(data_container, list):
        if clear_data_container:
            data_container.clear()
        for cell in sheet[axis]:
            val = value_former_transformer(cell.value)
            if value_filter(val):
                data_container.append(value_post_transformer(val))


def get_worksheet_data_by_column_title(
        data_container: set | list,
        clear_data_container: bool,
        sheet: ExcelWorksheet,
        column_title: str | None,
        value_former_transformer=lambda v: v,
        value_filter=lambda v: True,
        value_post_transformer=lambda v: v
) -> None:
    def find_column_index() -> int | None:
        if 1 < sheet.min_row or 1 > sheet.max_row:
            return None
        for column_i in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=column_i).value
            if (column_title is None and cell_value is None) or (column_title == cell_value):
                return column_i
        return None

    column_index = find_column_index()
    if column_index is None:
        return
    if isinstance(data_container, set):
        if clear_data_container:
            data_container.clear()
        for row_i in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row_i, column=column_index)
            val = value_former_transformer(cell.value)
            if value_filter(val):
                data_container.add(value_post_transformer(val))
    if isinstance(data_container, list):
        if clear_data_container:
            data_container.clear()
        for row_i in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row_i, column=column_index)
            val = value_former_transformer(cell.value)
            if value_filter(val):
                data_container.append(value_post_transformer(val))


def get_worksheet_data_by_row_title(
        data_container: set | list,
        clear_data_container: bool,
        sheet: ExcelWorksheet,
        row_title: str | None,
        value_former_transformer=lambda v: v,
        value_filter=lambda v: True,
        value_post_transformer=lambda v: v
) -> None:
    def find_row_index() -> int | None:
        if 1 < sheet.min_column or 1 > sheet.max_column:
            return None
        for row_i in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_i, column=1).value
            if (row_title is None and cell_value is None) or (row_title == cell_value):
                return row_i
        return None

    row_index = find_row_index()
    if row_index is None:
        return
    if isinstance(data_container, set):
        if clear_data_container:
            data_container.clear()
        for column_i in range(2, sheet.max_column + 1):
            cell = sheet.cell(row=row_index, column=column_i)
            val = value_former_transformer(cell.value)
            if value_filter(val):
                data_container.add(value_post_transformer(val))
    if isinstance(data_container, list):
        if clear_data_container:
            data_container.clear()
        for column_i in range(2, sheet.max_column + 1):
            cell = sheet.cell(row=row_index, column=column_i)
            val = value_former_transformer(cell.value)
            if value_filter(val):
                data_container.append(value_post_transformer(val))


def put_worksheet_data_by_axis(
        sheet: ExcelWorksheet,
        axis: str,
        value_title: str = "",
        print_item_count: bool = True,
        value_title_font: ExcelFont = ExcelFont(name="Times New Roman", bold=True, size=12),
        value_title_alignment: ExcelAlignment = ExcelAlignment(horizontal="center", vertical="center"),
        dataset: set | tuple | list | range = None,
        value_font: ExcelFont = ExcelFont(name="Times New Roman", color="000000", size=12),
        value_alignment: ExcelAlignment = ExcelAlignment(horizontal="center", vertical="center"),
        uniform_row_height: int = 20,
        uniform_column_width: int = 40
) -> None:
    if dataset is None:
        dataset = []
    if axis.isdigit():  # is non-negative integer, then represents a row
        row_number = int(axis)
        cell = sheet.cell(column=1, row=row_number)
        cell.value = value_title
        cell.font = value_title_font
        cell.alignment = value_title_alignment
        if print_item_count:
            cell = sheet.cell(column=2, row=row_number)
            cell.value = f"Item Count: {len(dataset)} (in this row)"
            cell.font = value_title_font
            cell.alignment = value_title_alignment
        for (i, string) in enumerate(dataset, start=(3 if print_item_count else 2)):
            cell = sheet.cell(column=i, row=row_number)
            cell.value = string
            cell.font = value_font
            cell.alignment = value_alignment
        sheet.row_dimensions[row_number].height = uniform_row_height
    else:  # is assumed to be capitalized letters, then represents a column
        cell = sheet[f"{axis}{1}"]
        cell.value = value_title
        cell.font = value_title_font
        cell.alignment = value_title_alignment
        if print_item_count:
            cell = sheet[f"{axis}{2}"]
            cell.value = f"Item Count: {len(dataset)} (in this column)"
            cell.font = value_title_font
            cell.alignment = value_title_alignment
        for (i, string) in enumerate(dataset, start=(3 if print_item_count else 2)):
            cell = sheet[f"{axis}{i}"]
            cell.value = string
            cell.font = value_font
            cell.alignment = value_alignment
        sheet.column_dimensions[axis].width = uniform_column_width
